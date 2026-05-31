"""Dump every sheet of Content architecture V.4.xlsx to a single text file."""
import openpyxl
from pathlib import Path

src = Path("Content architecture V.4.xlsx")
wb = openpyxl.load_workbook(src, data_only=True)

out_lines: list[str] = []
out_lines.append(f"# {src.name}")
out_lines.append(f"sheet names: {wb.sheetnames}")
out_lines.append("")

for sn in wb.sheetnames:
    ws = wb[sn]
    out_lines.append("=" * 80)
    out_lines.append(f"SHEET: {sn}  (rows={ws.max_row}, cols={ws.max_column})")
    out_lines.append("=" * 80)
    for row in ws.iter_rows(values_only=True):
        # Skip all-None rows for compactness
        if not any(c not in (None, "") for c in row):
            continue
        cells = [
            ("" if c is None else str(c).replace("\n", " | ").strip())
            for c in row
        ]
        out_lines.append("\t".join(cells))
    out_lines.append("")

Path("_content_architecture.txt").write_text("\n".join(out_lines), encoding="utf-8")
print(f"Wrote {len(out_lines)} lines from {len(wb.sheetnames)} sheets.")
