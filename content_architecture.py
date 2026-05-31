"""Content Architecture V.4 loader and matcher.

Single source of truth for the 50 pre-authored chatbot responses authored
by Iqbal Bagus Dwinanto (ESB Customer Support team). The Excel file
``Content architecture V.4.xlsx`` is the canonical content; this module
loads it at import time and exposes a keyword + category-filtered matcher
the agent uses to pick the right response for a merchant query.

Why we use the CA before falling back to Gemini synthesis:

* CA responses are reviewed and approved by Support before launch — no
  hallucination risk.
* They follow GILT (Greeting · Intro · Listing · Tail) and brand voice
  guidelines specific to ESB Order / OZE.
* They reference exact menu paths (``Master > Menu > Foto Menu``) and
  internal terminology (``Online Fund``, ``Push to POS``, ``Self Order
  Server``) the LLM might paraphrase incorrectly.

Public API
----------

``load_ca() -> list[dict]``
    Lazily parse the Excel file. Each dict has ``predefined``, ``prompt``,
    ``response``, ``tags``, ``keywords``, ``category``, ``entities``.

``match_ca(query, category=None, k=3) -> list[dict]``
    Keyword + token-overlap scoring; returns top-k entries, optionally
    filtered by an MVP category.

``format_response(text) -> str``
    Convert the CA pipe-delimited line breaks into plain-text newlines
    suitable for Telegram.
"""
from __future__ import annotations

import logging
import re
from functools import lru_cache
from pathlib import Path

logger = logging.getLogger(__name__)

CA_FILE = Path(__file__).with_name("Content architecture V.4.xlsx")

# Section headers in the workbook are like "01  AKTIVASI/NON AKTIVASI OZE".
# Each maps 1:1 to a PRD MVP category (English name used by classifier).
SECTION_TO_CATEGORY: dict[str, str] = {
    "01": "ESO Activation / Deactivation",
    "02": "Order Issues",
    "03": "Payment Gateway Setup",
    "04": "Menu Image Upload",
    "05": "Menu Issues",
    "06": "Banner Image Upload",
    "07": "ESO Merchant Issues",
    "08": "Payment & QR Issues",
    "09": "Guiding Configuration",
    "10": "Push to POS Issues",
}

MODEL_NAME = "Sukabantu 1.1"


def format_response(text: str) -> str:
    """Convert the CA's pipe-delimited line breaks to plain newlines.

    The authoring sheet uses `` | `` (space-pipe-space) as a line separator
    and ``|  |`` as a paragraph break. We normalize both to newlines so the
    response renders correctly in Telegram.
    """
    # First, paragraph break: " |  | " -> double newline.
    text = re.sub(r"\s*\|\s*\|\s*", "\n\n", text)
    # Remaining single pipes -> single newline.
    text = re.sub(r"\s*\|\s*", "\n", text)
    return text.strip()


@lru_cache(maxsize=1)
def load_ca() -> list[dict]:
    """Parse the Content Architecture workbook into a list of response entries.

    Returns ``[]`` if the file is missing — callers should treat this as
    "fall back to Gemini synthesis." Logged loudly so it's discoverable.
    """
    if not CA_FILE.exists():
        logger.warning("Content Architecture file not found at %s", CA_FILE)
        return []

    try:
        import openpyxl  # local import keeps this optional at install time
    except ImportError:
        logger.warning("openpyxl not installed; cannot load Content Architecture.")
        return []

    wb = openpyxl.load_workbook(CA_FILE, data_only=True)
    sheet_name = "Chatbot content architecture"
    if sheet_name not in wb.sheetnames:
        logger.warning("Sheet %r not found in %s", sheet_name, CA_FILE.name)
        return []
    ws = wb[sheet_name]

    entries: list[dict] = []
    current_category: str | None = None
    section_re = re.compile(r"^(\d{1,2})\s+(.+)$")

    # Header rows occupy the top of the sheet (title + counts + column names).
    # Data starts where rows have a non-empty AI Response in column D (index 3).
    for row in ws.iter_rows(values_only=True):
        if not row or all(c in (None, "") for c in row):
            continue

        first_cell = "" if row[0] is None else str(row[0]).strip()

        # Section header row: first cell like "01  AKTIVASI..." and rest blank.
        m = section_re.match(first_cell)
        if m and all(c in (None, "") for c in row[1:]):
            num = m.group(1).zfill(2)
            current_category = SECTION_TO_CATEGORY.get(num)
            if current_category is None:
                logger.warning("Unknown CA section %r — skipping", first_cell)
            continue

        # Data row: must have an AI Response in column D.
        if len(row) < 4 or not row[3]:
            continue
        if current_category is None:
            continue  # data row before any section header

        # Cells use either "|" or newlines as separators (authoring varies);
        # split on both.
        keywords_raw = str(row[5] or "")
        keywords = [
            kw.lstrip("•").strip()
            for kw in re.split(r"[\n|]+", keywords_raw)
            if kw.strip()
        ]
        tags_raw = str(row[4] or "")
        tags = [t.strip() for t in re.split(r"[\n|]+", tags_raw) if t.strip()]

        entries.append({
            "predefined": first_cell,
            "prompt": str(row[1] or "").strip(),
            "model": str(row[2] or MODEL_NAME).strip(),
            "response": str(row[3] or "").strip(),
            "tags": tags,
            "keywords": keywords,
            "structure": str(row[6] or "").strip() if len(row) > 6 else "",
            "entities": str(row[7] or "").strip() if len(row) > 7 else "",
            "category": current_category,
        })

    logger.info("Loaded %d Content Architecture responses across %d categories.",
                len(entries), len({e["category"] for e in entries}))
    return entries


# Indonesian stopwords we ignore when token-matching to reduce noise.
_STOPWORDS = {
    "yang", "untuk", "dari", "saya", "ada", "tidak", "bisa", "saat", "atau",
    "dan", "atau", "ini", "itu", "di", "ke", "dengan", "pada", "sudah",
    "akan", "bagaimana", "cara", "kenapa", "kalau", "jika", "apa", "tapi",
    "hanya", "agar", "bila", "saja",
}


# ESB-specific synonym pairs. Merchants and CA authors use different words
# for the same thing — we normalize both into a canonical form before matching
# so e.g. "foto menu tidak muncul" matches keyword "gambar menu tidak muncul".
_SYNONYMS: list[tuple[str, str]] = [
    (r"\bfoto\b", "gambar"),       # foto ↔ gambar (menu image)
    (r"\beso\b", "oze"),           # ESO was renamed to OZE
    (r"\bgagal\b", "tidak bisa"),  # gagal ↔ tidak bisa
    (r"\berror\b", "kendala"),
    (r"\bpic\b", "gambar"),
]


def _normalize(text: str) -> str:
    n = re.sub(r"\s+", " ", text.lower().strip())
    for pat, repl in _SYNONYMS:
        n = re.sub(pat, repl, n)
    return n


def _tokens(text: str) -> set[str]:
    norm = _normalize(text)
    return {t for t in re.findall(r"[a-z0-9]+", norm) if len(t) > 2 and t not in _STOPWORDS}


def match_ca(query: str, category: str | None = None, k: int = 3) -> list[dict]:
    """Return the top-k CA entries most relevant to the query.

    Scoring (deterministic, no LLM call):

    * +3.0 for each keyword in the entry that appears as a substring of
      the query (these are curated trigger phrases — strong signal).
    * +1.0 for each predefined-key token that overlaps with the query.
    * +0.5 for each prompt token that overlaps with the query.
    * +1.5 bonus when the entry's category matches the classifier's pick.

    Returns ``[]`` if nothing scores above zero.
    """
    entries = load_ca()
    if not entries:
        return []

    q = _normalize(query)
    q_tokens = _tokens(query)

    scored: list[tuple[float, dict]] = []
    for e in entries:
        score = 0.0
        for kw in e["keywords"]:
            if not kw:
                continue
            if _normalize(kw) in q:
                score += 3.0
        for tok in _tokens(e["predefined"]):
            if tok in q_tokens:
                score += 1.0
        for tok in _tokens(e["prompt"]):
            if tok in q_tokens:
                score += 0.5
        if category and e["category"] == category:
            score += 1.5
        if score > 0:
            scored.append((score, e))

    scored.sort(key=lambda x: x[0], reverse=True)
    return [{**e, "match_score": round(s, 2)} for s, e in scored[:k]]


def list_categories() -> list[str]:
    """All MVP categories represented in the CA — useful for taxonomy validation."""
    return sorted({e["category"] for e in load_ca()})


def entries_in_category(category: str) -> list[dict]:
    """All CA entries belonging to a given PRD category, in authoring order."""
    return [e for e in load_ca() if e["category"] == category]


def find_by_predefined(predefined: str) -> dict | None:
    """Look up a single CA entry by its predefined key (exact match)."""
    needle = predefined.strip()
    for e in load_ca():
        if e["predefined"] == needle:
            return e
    return None
