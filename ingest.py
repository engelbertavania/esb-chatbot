"""Convert the categorized ticket Excel into a JSONL corpus for retrieval.

Output: ``vertex_corpus.jsonl`` (one JSON object per line).

Each row in the Excel becomes one document with this shape (compatible with
both Vertex AI Search structured datastores and our local fallback retriever
in ``rag.py``)::

    {
      "id": "ticket-0001",
      "content": "Issue: ...\\nDetail: ...",
      "metadata": {
        "category_tag": "order_issue",
        "tags": "Ordering Process",
        "services": "Question",
        "priority": "Low",
        "issue": "Kendala Pesanan",
        "detail": "...",
        "root_cause": "...",
        "solution": "..."
      }
    }

To load into Vertex AI Search:

  1. Upload ``vertex_corpus.jsonl`` to a GCS bucket.
  2. In the GCP console, create a Vertex AI Search **data store** (generic,
     structured data, JSONL).
  3. Import the file (``content`` becomes the indexed text; ``metadata.*``
     becomes filterable struct fields).
  4. Set the ``VERTEX_DATASTORE_ID`` env var (and ``GOOGLE_CLOUD_PROJECT`` /
     ``GOOGLE_CLOUD_LOCATION``) before starting the backend.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

EXCEL_PATH = Path(__file__).parent / "Ticket_Order_Test_Q1_2026_Categorized_v2.xlsx"
OUTPUT_PATH = Path(__file__).parent / "vertex_corpus.jsonl"


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_corpus(excel_path: Path = EXCEL_PATH, output_path: Path = OUTPUT_PATH) -> int:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h else "" for h in next(rows)]
    col = {name: idx for idx, name in enumerate(header)}

    required = ["Issue", "Detail", "Root Cause", "Solution", "Category Tag",
                "Tags", "Services", "Priority"]
    missing = [c for c in required if c not in col]
    if missing:
        raise SystemExit(f"Excel is missing expected columns: {missing}")

    written = 0
    skipped = 0
    with output_path.open("w", encoding="utf-8") as f:
        for i, row in enumerate(rows, start=1):
            issue = _clean(row[col["Issue"]])
            detail = _clean(row[col["Detail"]])
            if not issue and not detail:
                skipped += 1
                continue

            doc = {
                "id": f"ticket-{i:04d}",
                "content": f"Issue: {issue}\nDetail: {detail}".strip(),
                "metadata": {
                    "category_tag": _clean(row[col["Category Tag"]]),
                    "tags": _clean(row[col["Tags"]]),
                    "services": _clean(row[col["Services"]]),
                    "priority": _clean(row[col["Priority"]]),
                    "issue": issue,
                    "detail": detail,
                    "root_cause": _clean(row[col["Root Cause"]]),
                    "solution": _clean(row[col["Solution"]]),
                },
            }
            f.write(json.dumps(doc, ensure_ascii=False) + "\n")
            written += 1

    return written


if __name__ == "__main__":
    n = build_corpus()
    print(f"Wrote {n} documents to {OUTPUT_PATH}")
    if n == 0:
        sys.exit(1)
